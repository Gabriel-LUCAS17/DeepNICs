import os
import sys
os.chdir(sys.path[0])

from src.CountRes import getVarMatrix
from openpyxl import load_workbook,Workbook
from multiprocessing import Pool
import subprocess
import glob

colNames = ['Var','Nb tirages', 'Se,Sp = 1', 'Coeffs != 0', 'Coeffs > 0',
            'Coeffs < 0', 'Autres cas', 'Nb cycles = 1', 'Nb cycles = 1 et Nb etapes = 1',
            'Toutes solutions != 4', 'Toutes solutions = 1']

def countResults(var,folderPath, resName, AP_THRESHOLD):
    """
    Generates a .csv file summarizing the results obtained for a ROP-Forest.

    Parameters :
        folderPath : string. Path of folder containing the ROP-Forest.
        resName : string. Path of .csv file to write.

    Returns :
        no return.
    """
    global colNames
    nbCol = len(colNames)

    # Getting the matrix which will be written on disk.
    varMatrix = getVarMatrix(folderPath, AP_THRESHOLD,nbCol)

    if os.path.exists(resName):
        wb = load_workbook(resName)
    else:
        wb = Workbook()
    
    ws = wb.active

    for i in range(len(colNames)):
        ws.cell(row=1,column=i+1).value = colNames[i]
    ws.cell(row=var+1,column=1).value = var

    #print("Processing : ", folderPath.split("/")[-3:])

    for i in range(len(varMatrix)):
        ws.cell(row=var+1,column=i+2).value = varMatrix[i]
    # Write varMatrix content on disk.
    wb.save(resName)

def countVarResults(list_var, R, V, data_path, resPath, AP_THRESHOLD):
    """
    Calls countResults function for given stratified variable for all configurations.

    Parameters :
        COULD CHANGE.

    Returns :
        no return.
    """

    # Defining reference path and result path.

    # Summing up results for all configurations of range of simulation and nb of selected variables.
    for var in list_var:
        refPath = data_path + "a" + str(var) + "/"
        folderPath = refPath + "R" + str(R) + "/V" + str(V)
        resName = resPath + "R" + str(R) + "V" + str(V) + "_res.xlsx"
        countResults(var,folderPath, resName, AP_THRESHOLD)

if __name__ == '__main__':

    def Usage() :
        """
        Prints the Usage of ROP-Image-Generation.py.

        Parameters :
            no parameters.

        Return :
            no return.
        """
        print("\nUsage of ROP-Image-Generation.py : ")
        print("\tThis script must be executed from the Anaconda Prompt.")
        print("\tSeveral command-line arguments must be specified.\n")
        print("Usage : 'python CountRes.py arg1 arg2 arg3 arg4'")
        print("where :")
        print("\t\t arg1 is the minimum range of simulation range (integer)")
        print("\t\t arg2 is the maximum range of simulation range (integer)")
        print("\t\t arg3 is the minimum number of selected variables (integer)")
        print("\t\t arg4 is the maximum number of selected variables (integer)")
        exit()

    def Execute_Command(command) :
        """
        Execute specified command using subprocess function.

        Parameters :
            command : string or list of strings. Command to execute.

        Returns :
            no return.
        """

        program = subprocess.run(command)
        if program.returncode == 0 :
            print("Program correctly executed.")
        else :
            print("Failure to execute this program.")
            print("Command was : ", command)
            print(repr(program.stdout))
            print(repr(program.stderr))
            exit()

    def check_int(x) :
        """
        Returns True if the specified argument is an integer, False otherwise.

        Parameters :
            x : string.

        Returns :
            is_integer : bool.
        """
        try :
            int(x)
            return True
        except :
            return False

    def getListVariables(path) :
        """
        Returns list of variables contained in the specified folder.

        Parameters :
            path : string. Path for folder containing data for several stratified
                variable.

        Returns :
            list_var : list of int. List of variables' id contained in the
                specified folder.
        """

        list_var = []
        pattern = path + "a*"
        list_folders = glob.glob(pattern)
        for folder in list_folders :
            list_var += [int(folder.split("\\")[-1][1:])]
        list_var.sort()
        return list_var

    # Checking number of arguments and arguments' consistency.
    if len(sys.argv) == 5 :
        print("\nList of entered arguments : ", sys.argv[1:])
        print("Checking arguments' consistency.")
        # Are the entered args recognized as integers ?
        for val in sys.argv[1:] :
            if not check_int(val) :
                print("ERROR: the argument ", val, " is not an integer.")
                Usage()
        # Have we : infR < supR and infV < supV ?
        if int(sys.argv[1]) > int(sys.argv[2]) or int(sys.argv[3]) > int(sys.argv[4]) :
            print("ERROR: Entered values are not consistent.")
            Usage()
        else :
            print("Entered arguments are consistent.")
            infR = int(sys.argv[1])
            supR = int(sys.argv[2])
            infV = int(sys.argv[3])
            supV = int(sys.argv[4])
            width  = supR - infR + 1
            height = supV - infV + 1
    else :
        print("\nWrong number of arguments. 4 arguments expected.")
        Usage()

    # Get constants' definition from command-line.
    data_path = str(input("Enter path to folder containing all ROP-Training samples: "))
    res_path = str(input("Enter path to result folder"))
    AP_THRESHOLD = 1
    N_proc = os.cpu_count()
    print("Number of processors to use set to : ", N_proc)

    if not os.path.exists(res_path):
        os.mkdir(res_path)

    list_var = getListVariables(data_path)
    print(data_path)
    print(list_var)

    List_Args_Count_Res = [(list_var, i, j, data_path, res_path, AP_THRESHOLD) for i in range(infR,supR+1) for j in range(infV,supV+1)]

    with Pool(N_proc) as p:
        print("Counting results.")
        p.starmap(countVarResults, List_Args_Count_Res)